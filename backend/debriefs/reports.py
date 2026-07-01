"""
Report generation helpers.
Supports Excel (openpyxl) and PDF (ReportLab) exports.
"""
import io
from datetime import date

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
)

from .models import DailyDebrief, HourlyLog


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _get_report_data(start_date: date, end_date: date, intern_id=None):
    """Return queryset-based data for the date range."""
    debrief_qs = DailyDebrief.objects.filter(
        date__gte=start_date, date__lte=end_date
    ).select_related('intern', 'feedback', 'feedback__supervisor').order_by('intern__username', 'date')

    log_qs = HourlyLog.objects.filter(
        date__gte=start_date, date__lte=end_date
    ).select_related('intern').order_by('intern__username', 'date', 'start_time')

    if intern_id:
        debrief_qs = debrief_qs.filter(intern_id=intern_id)
        log_qs = log_qs.filter(intern_id=intern_id)

    return list(debrief_qs), list(log_qs)


# ─── Excel ────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")


def _style_header(ws, row, cols):
    for col_idx, _ in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def generate_excel_report(start_date: date, end_date: date, intern_id=None) -> HttpResponse:
    debriefs, logs = _get_report_data(start_date, end_date, intern_id)

    wb = Workbook()

    # ── Sheet 1: Daily Debriefs ──
    ws1 = wb.active
    ws1.title = "Daily Debriefs"
    debrief_headers = [
        'Intern', 'Date', 'Yesterday Task', 'Progress Made',
        'Challenges', 'Today Task', 'Notes', 'Feedback', 'Feedback By',
    ]
    ws1.append(debrief_headers)
    _style_header(ws1, 1, debrief_headers)

    for i, d in enumerate(debriefs, 2):
        feedback_content = d.feedback.content if hasattr(d, 'feedback') and d.feedback else ''
        feedback_by = (
            d.feedback.supervisor.get_full_name() or d.feedback.supervisor.username
            if hasattr(d, 'feedback') and d.feedback and d.feedback.supervisor
            else ''
        )
        row = [
            d.intern.get_full_name() or d.intern.username,
            str(d.date),
            d.yesterday_task,
            d.progress_made,
            d.challenges,
            d.today_task,
            d.notes or '',
            feedback_content,
            feedback_by,
        ]
        ws1.append(row)
        if i % 2 == 0:
            for col_idx in range(1, len(debrief_headers) + 1):
                ws1.cell(row=i, column=col_idx).fill = ALT_FILL

    # Auto-width
    for col in ws1.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # ── Sheet 2: Hourly Logs ──
    ws2 = wb.create_sheet("Hourly Logs")
    log_headers = [
        'Intern', 'Date', 'Start Time', 'End Time',
        'Activity', 'Productivity Score (1-5)', 'Duration (hrs)',
    ]
    ws2.append(log_headers)
    _style_header(ws2, 1, log_headers)

    for i, log in enumerate(logs, 2):
        row = [
            log.intern.get_full_name() or log.intern.username,
            str(log.date),
            str(log.start_time),
            str(log.end_time),
            log.activity,
            log.productivity_score,
            log.duration_hours,
        ]
        ws2.append(row)
        if i % 2 == 0:
            for col_idx in range(1, len(log_headers) + 1):
                ws2.cell(row=i, column=col_idx).fill = ALT_FILL

    for col in ws2.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # ── Response ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"DebriefPro_Report_{start_date}_to_{end_date}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── PDF ──────────────────────────────────────────────────────────────────────

def generate_pdf_report(start_date: date, end_date: date, intern_id=None) -> HttpResponse:
    debriefs, logs = _get_report_data(start_date, end_date, intern_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        textColor=colors.HexColor('#1E40AF'),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        textColor=colors.grey,
        spaceAfter=12,
    )
    heading2 = ParagraphStyle(
        'Heading2Custom', parent=styles['Heading2'],
        textColor=colors.HexColor('#1E40AF'),
        spaceBefore=14,
        spaceAfter=6,
    )
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10)

    story = []

    # Title
    story.append(Paragraph("DebriefPro — Internship Report", title_style))
    story.append(Paragraph(
        f"Period: {start_date.strftime('%B %d, %Y')} – {end_date.strftime('%B %d, %Y')}",
        subtitle_style,
    ))
    story.append(Spacer(1, 0.3 * cm))

    header_color = colors.HexColor('#1E40AF')
    alt_color = colors.HexColor('#EFF6FF')
    white = colors.white

    def build_table(headers, rows, col_widths):
        data = [[Paragraph(f'<b>{h}</b>', ParagraphStyle('TH', parent=cell_style,
                                                          textColor=white, fontSize=8)) for h in headers]]
        for row in rows:
            data.append([Paragraph(str(v), cell_style) for v in row])

        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, alt_color]),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('PADDING', (0, 0), (-1, -1), 4),
        ])
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(style)
        return t

    # ── Daily Debriefs ──
    story.append(Paragraph("Daily Debriefs", heading2))
    if debriefs:
        debrief_headers = ['Intern', 'Date', 'Yesterday Task', 'Progress', 'Challenges', 'Today Task', 'Feedback']
        debrief_rows = []
        for d in debriefs:
            fb = d.feedback.content[:120] + '…' if hasattr(d, 'feedback') and d.feedback and len(d.feedback.content) > 120 else (d.feedback.content if hasattr(d, 'feedback') and d.feedback else '—')
            debrief_rows.append([
                d.intern.get_full_name() or d.intern.username,
                str(d.date),
                d.yesterday_task[:100],
                d.progress_made[:100],
                d.challenges[:100],
                d.today_task[:100],
                fb,
            ])
        col_w = [3 * cm, 2.2 * cm, 4 * cm, 4 * cm, 4 * cm, 4 * cm, 4.5 * cm]
        story.append(build_table(debrief_headers, debrief_rows, col_w))
    else:
        story.append(Paragraph("No debrief records found for this period.", styles['Normal']))

    story.append(Spacer(1, 0.5 * cm))

    # ── Hourly Logs ──
    story.append(Paragraph("Hourly Activity Logs", heading2))
    if logs:
        log_headers = ['Intern', 'Date', 'Start', 'End', 'Activity', 'Score', 'Hrs']
        log_rows = [
            [
                log.intern.get_full_name() or log.intern.username,
                str(log.date),
                str(log.start_time),
                str(log.end_time),
                log.activity[:140],
                log.productivity_score,
                log.duration_hours,
            ]
            for log in logs
        ]
        col_w = [3 * cm, 2.2 * cm, 1.8 * cm, 1.8 * cm, 12 * cm, 1.5 * cm, 1.5 * cm]
        story.append(build_table(log_headers, log_rows, col_w))
    else:
        story.append(Paragraph("No hourly log records found for this period.", styles['Normal']))

    doc.build(story)
    buffer.seek(0)

    filename = f"DebriefPro_Report_{start_date}_to_{end_date}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
